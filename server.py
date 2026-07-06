"""加班登记系统 - HTTP 服务入口

用法:
    python server.py sqlite3 [--db-path PATH] [--host HOST] [--port PORT]
    python server.py mysql   --host H --port P --user U --password PWD --database DB [--host HOST] [--port PORT]
    python server.py pg     --host H --port P --user U --password PWD --database DB [--host HOST] [--port PORT]

别名: db=sqlite3, my=mysql, pg=postgresql
"""

import argparse
import io
import json
import os
import sys
import urllib.parse
from http.server import HTTPServer, SimpleHTTPRequestHandler
from datetime import datetime

# 将项目目录加入 sys.path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import config
import db

# ---- 子命令别名映射 ----
_ALIAS_MAP = {
    "db": "sqlite3",
    "sqlite": "sqlite3",
    "sqlite3": "sqlite3",
    "my": "mysql",
    "mysql": "mysql",
    "pg": "postgresql",
    "postgresql": "postgresql",
}


def build_parser():
    """构建子命令解析器"""
    parser = argparse.ArgumentParser(
        prog="server.py",
        description="加班登记系统 - 支持子命令选择数据库类型",
        epilog="示例:\n"
               "  python server.py sqlite3                    # 使用默认 SQLite\n"
               "  python server.py sqlite3 --db-path data.db  # 指定 SQLite 路径\n"
               "  python server.py mysql --db-host 127.0.0.1 --db-user root --db-password 123 --db-database ot --host 0.0.0.0 --port 8080\n"
               "  python server.py pg --db-host 127.0.0.1 --db-user postgres --db-password abc --db-database ot --host 0.0.0.0 --port 8080\n",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )

    subparsers = parser.add_subparsers(dest="db_type", help="数据库类型 (三选一)")

    # 公共参数
    def add_common_args(sub_p):
        sub_p.add_argument("--host", default="0.0.0.0", help="监听地址 (默认: 0.0.0.0)")
        sub_p.add_argument("--port", type=int, default=8080, help="监听端口 (默认: 8080)")

    # ---- sqlite3 / db / sqlite ----
    p_sqlite = subparsers.add_parser(
        "sqlite3", aliases=["db", "sqlite"],
        help="使用 SQLite 数据库 (零配置，默认方式)",
    )
    p_sqlite.add_argument("--db-path", default=None, help="SQLite 数据库文件路径 (默认: overtime.db)")
    add_common_args(p_sqlite)

    # ---- mysql / my ----
    p_mysql = subparsers.add_parser(
        "mysql", aliases=["my"],
        help="使用 MySQL 数据库",
    )
    p_mysql.add_argument("--db-host", required=True, help="MySQL 主机地址")
    p_mysql.add_argument("--db-port", type=int, default=None, help="MySQL 端口 (默认: 3306)")
    p_mysql.add_argument("--db-user", required=True, help="MySQL 用户名")
    p_mysql.add_argument("--db-password", default="", help="MySQL 密码 (默认: 空)")
    p_mysql.add_argument("--db-database", required=True, help="MySQL 数据库名")
    p_mysql.add_argument("--db-charset", default="utf8mb4", help="MySQL 字符集 (默认: utf8mb4)")
    add_common_args(p_mysql)

    # ---- postgresql / pg ----
    p_pg = subparsers.add_parser(
        "postgresql", aliases=["pg"],
        help="使用 PostgreSQL 数据库",
    )
    p_pg.add_argument("--db-host", required=True, help="PostgreSQL 主机地址")
    p_pg.add_argument("--db-port", type=int, default=None, help="PostgreSQL 端口 (默认: 5432)")
    p_pg.add_argument("--db-user", required=True, help="PostgreSQL 用户名")
    p_pg.add_argument("--db-password", default="", help="PostgreSQL 密码 (默认: 空)")
    p_pg.add_argument("--db-database", required=True, help="PostgreSQL 数据库名")
    add_common_args(p_pg)

    return parser


class OvertimeHandler(SimpleHTTPRequestHandler):
    """加班登记系统请求处理器"""

    def __init__(self, *args, **kwargs):
        self.server_dir = os.path.dirname(os.path.abspath(__file__))
        super().__init__(*args, directory=os.path.join(self.server_dir, "static"), **kwargs)

    def log_message(self, format, *args):
        """自定义日志格式"""
        print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {args[0]}")

    def _send_json(self, data, status=200):
        """发送 JSON 响应"""
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.end_headers()
        self.wfile.write(json.dumps(data, ensure_ascii=False, default=str).encode("utf-8"))

    def _send_cors_headers(self):
        """发送 CORS 头"""
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, PUT, DELETE, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")

    def _read_body(self):
        """读取请求体"""
        content_length = int(self.headers.get("Content-Length", 0))
        return self.rfile.read(content_length).decode("utf-8")

    def do_OPTIONS(self):
        """处理 OPTIONS 预检请求"""
        self.send_response(200)
        self._send_cors_headers()
        self.end_headers()

    def do_GET(self):
        """GET 请求路由"""
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path

        if path == "/api/records":
            self._handle_list_records(parsed.query)
        elif path == "/api/records/export":
            self._handle_export(parsed.query)
        elif path == "/api/auto-fill":
            self._handle_auto_fill(parsed.query)
        elif path == "/api/stats":
            self._handle_stats()
        elif path.startswith("/api/records/") and path.count("/") == 3:
            record_id = path.split("/")[-1]
            self._handle_get_record(record_id)
        elif path.startswith("/api/"):
            self._send_json({"error": "接口不存在"}, 404)
        else:
            # 静态文件由 SimpleHTTPRequestHandler 处理
            super().do_GET()

    def do_POST(self):
        """POST 请求路由"""
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path

        if path == "/api/records":
            self._handle_create_record()
        elif path.startswith("/api/"):
            self._send_json({"error": "接口不存在"}, 404)
        else:
            self._send_json({"error": "无效的请求路径"}, 404)

    def do_PUT(self):
        """PUT 请求路由"""
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path

        if path.startswith("/api/records/"):
            record_id = path.split("/")[-1]
            self._handle_update_record(record_id)
        else:
            self._send_json({"error": "无效的请求路径"}, 404)

    def do_DELETE(self):
        """DELETE 请求路由"""
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path

        if path.startswith("/api/records/"):
            record_id = path.split("/")[-1]
            self._handle_delete_record(record_id)
        else:
            self._send_json({"error": "无效的请求路径"}, 404)

    # ---- API 处理方法 ----

    def _handle_create_record(self):
        """处理创建加班记录"""
        try:
            body = self._read_body()
            data = json.loads(body)
        except (json.JSONDecodeError, ValueError):
            self._send_json({"error": "请求格式错误，请使用JSON格式"}, 400)
            return

        # 字段校验
        required_fields = [
            "employee_name", "employee_id", "company",
            "project_team", "pm", "reason",
            "overtime_date", "start_time", "end_time",
        ]
        missing = [f for f in required_fields if not data.get(f, "").strip()]
        if missing:
            self._send_json({"error": f"以下字段为必填项: {', '.join(missing)}"}, 400)
            return

        # 日期格式校验
        for field_name in ["overtime_date"]:
            val = data[field_name].strip()
            try:
                datetime.strptime(val, "%Y-%m-%d")
            except ValueError:
                self._send_json({"error": f"{field_name} 格式错误，应为 yyyy-mm-dd"}, 400)
                return

        # 时间格式校验
        for field_name in ["start_time", "end_time"]:
            val = data[field_name].strip()
            try:
                datetime.strptime(val, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                self._send_json({"error": f"{field_name} 格式错误，应为 yyyy-mm-dd hh24:mi:ss"}, 400)
                return

        # 结束时间校验
        if data["end_time"] <= data["start_time"]:
            self._send_json({"error": "结束时间必须晚于开始时间"}, 400)
            return

        try:
            record_id = db.insert_record({k: v.strip() for k, v in data.items()})
            self._send_json({"message": "登记成功", "id": record_id}, 201)
        except Exception as e:
            err_msg = str(e)
            if "UNIQUE" in err_msg or "Duplicate" in err_msg:
                self._send_json({"error": "该工号在此日期已存在登记记录"}, 409)
            else:
                self._send_json({"error": f"服务器错误: {err_msg}"}, 500)

    def _handle_update_record(self, record_id):
        """处理更新加班记录"""
        try:
            record_id = int(record_id)
        except ValueError:
            self._send_json({"error": "无效的记录ID"}, 400)
            return

        existing = db.get_record(record_id)
        if not existing:
            self._send_json({"error": "记录不存在"}, 404)
            return

        try:
            body = self._read_body()
            data = json.loads(body)
        except (json.JSONDecodeError, ValueError):
            self._send_json({"error": "请求格式错误"}, 400)
            return

        required_fields = [
            "employee_name", "employee_id", "company",
            "project_team", "pm", "reason",
            "overtime_date", "start_time", "end_time",
        ]
        missing = [f for f in required_fields if not data.get(f, "").strip()]
        if missing:
            self._send_json({"error": f"以下字段为必填项: {', '.join(missing)}"}, 400)
            return

        for field_name in ["overtime_date"]:
            try:
                datetime.strptime(data[field_name].strip(), "%Y-%m-%d")
            except ValueError:
                self._send_json({"error": f"{field_name} 格式错误，应为 yyyy-mm-dd"}, 400)
                return

        for field_name in ["start_time", "end_time"]:
            try:
                datetime.strptime(data[field_name].strip(), "%Y-%m-%d %H:%M:%S")
            except ValueError:
                self._send_json({"error": f"{field_name} 格式错误"}, 400)
                return

        if data["end_time"] <= data["start_time"]:
            self._send_json({"error": "结束时间必须晚于开始时间"}, 400)
            return

        try:
            db.update_record(record_id, {k: v.strip() for k, v in data.items()})
            self._send_json({"message": "更新成功"})
        except Exception as e:
            err_msg = str(e)
            if "UNIQUE" in err_msg or "Duplicate" in err_msg:
                self._send_json({"error": "该工号在此日期已存在登记记录"}, 409)
            else:
                self._send_json({"error": f"服务器错误: {err_msg}"}, 500)

    def _handle_delete_record(self, record_id):
        """处理删除加班记录"""
        try:
            record_id = int(record_id)
        except ValueError:
            self._send_json({"error": "无效的记录ID"}, 400)
            return

        existing = db.get_record(record_id)
        if not existing:
            self._send_json({"error": "记录不存在"}, 404)
            return

        db.delete_record(record_id)
        self._send_json({"message": "删除成功"})

    def _handle_list_records(self, query_string):
        """处理查询加班记录列表"""
        params = urllib.parse.parse_qs(query_string)

        date_from = params.get("date_from", [None])[0]
        date_to = params.get("date_to", [None])[0]
        company = params.get("company", [None])[0]
        project_team = params.get("project_team", [None])[0]
        employee_id = params.get("employee_id", [None])[0]
        employee_name = params.get("employee_name", [None])[0]
        page = int(params.get("page", [1])[0])
        page_size = int(params.get("page_size", [50])[0])

        result = db.query_records(
            date_from=date_from, date_to=date_to,
            company=company, project_team=project_team,
            employee_id=employee_id, employee_name=employee_name,
            page=page, page_size=page_size,
        )
        self._send_json(result)

    def _handle_auto_fill(self, query_string):
        """自动填充：根据姓名或工号查找最近登记的关联信息"""
        params = urllib.parse.parse_qs(query_string)
        name = params.get("name", [None])[0]
        employee_id = params.get("employee_id", [None])[0]

        if not name and not employee_id:
            self._send_json({"error": "请提供姓名或工号"}, 400)
            return

        result = db.lookup_employee_info(name=name, employee_id=employee_id)
        self._send_json(result)

    def _handle_stats(self):
        """获取概览统计数据"""
        result = db.get_stats()
        self._send_json(result)

    def _handle_get_record(self, record_id):
        """获取单条记录"""
        try:
            record_id = int(record_id)
        except ValueError:
            self._send_json({"error": "无效的记录ID"}, 400)
            return

        record = db.get_record(record_id)
        if not record:
            self._send_json({"error": "记录不存在"}, 404)
            return
        self._send_json(record)

    def _handle_export(self, query_string):
        """处理Excel导出"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        except ImportError:
            self._send_json({"error": "服务端未安装 openpyxl，请执行: pip install openpyxl"}, 500)
            return

        params = urllib.parse.parse_qs(query_string)
        date_from = params.get("date_from", [None])[0]
        date_to = params.get("date_to", [None])[0]
        company = params.get("company", [None])[0]
        project_team = params.get("project_team", [None])[0]
        employee_id = params.get("employee_id", [None])[0]
        employee_name = params.get("employee_name", [None])[0]

        records = db.export_all_records(
            date_from=date_from, date_to=date_to,
            company=company, project_team=project_team,
            employee_id=employee_id, employee_name=employee_name,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "加班登记表"

        # 表头样式
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        headers = ["序号", "加班人员", "工号", "合作公司", "项目组", "PM",
                    "加班事由", "计划加班日期", "计划开始时间", "计划结束时间", "登记时间"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # 数据行
        for row_idx, record in enumerate(records, 2):
            values = [
                row_idx - 1,
                record.get("employee_name", ""),
                record.get("employee_id", ""),
                record.get("company", ""),
                record.get("project_team", ""),
                record.get("pm", ""),
                record.get("reason", ""),
                record.get("overtime_date", ""),
                record.get("start_time", ""),
                record.get("end_time", ""),
                record.get("created_at", ""),
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # 列宽
        col_widths = [6, 12, 12, 20, 20, 12, 30, 14, 22, 22, 22]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = width

        # 冻结首行
        ws.freeze_panes = "A2"

        # 输出到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"加班登记表_{datetime.now().strftime('%Y%m%d')}.xlsx"
        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{urllib.parse.quote(filename)}")
        self.send_header("Content-Length", str(len(output.getvalue())))
        self.end_headers()
        self.wfile.write(output.getvalue())


def main():
    """主函数 - 子命令方式启动"""
    parser = build_parser()

    # 不足 2 个参数时打印帮助
    if len(sys.argv) < 2 or sys.argv[1] in ("-h", "--help"):
        parser.print_help()
        sys.exit(0)

    args = parser.parse_args()
    db_type = _ALIAS_MAP.get(args.db_type, args.db_type) if hasattr(args, "db_type") else args.db_type

    # 根据子命令设置配置
    if db_type == "sqlite3":
        config.DB_TYPE = "sqlite"
        if args.db_path:
            config.SQLITE_DB_PATH = args.db_path
        print(f"[INFO] 数据库类型: SQLite")
        print(f"[INFO] 数据库文件: {config.SQLITE_DB_PATH}")

    elif db_type == "mysql":
        config.DB_TYPE = "mysql"
        config.MYSQL_CONFIG["host"] = args.db_host
        config.MYSQL_CONFIG["port"] = args.db_port or 3306
        config.MYSQL_CONFIG["user"] = args.db_user
        config.MYSQL_CONFIG["password"] = args.db_password
        config.MYSQL_CONFIG["database"] = args.db_database
        config.MYSQL_CONFIG["charset"] = args.db_charset
        print(f"[INFO] 数据库类型: MySQL")
        print(f"[INFO] 连接地址: {config.MYSQL_CONFIG['host']}:{config.MYSQL_CONFIG['port']}/{config.MYSQL_CONFIG['database']}")
        print(f"[INFO] 连接用户: {config.MYSQL_CONFIG['user']}")

    elif db_type == "postgresql":
        config.DB_TYPE = "postgresql"
        config.POSTGRESQL_CONFIG["host"] = args.db_host
        config.POSTGRESQL_CONFIG["port"] = args.db_port or 5432
        config.POSTGRESQL_CONFIG["user"] = args.db_user
        config.POSTGRESQL_CONFIG["password"] = args.db_password
        config.POSTGRESQL_CONFIG["database"] = args.db_database
        print(f"[INFO] 数据库类型: PostgreSQL")
        print(f"[INFO] 连接地址: {config.POSTGRESQL_CONFIG['host']}:{config.POSTGRESQL_CONFIG['port']}/{config.POSTGRESQL_CONFIG['database']}")
        print(f"[INFO] 连接用户: {config.POSTGRESQL_CONFIG['user']}")

    print(f"[INFO] 监听地址: {args.host}:{args.port}")
    print("")

    # 初始化数据库
    db.init_db()
    print("[INFO] 数据库初始化完成")
    print("")

    # 启动服务
    server = HTTPServer((args.host, args.port), OvertimeHandler)
    print(f"[INFO] 服务已启动: http://{args.host}:{args.port}")
    print("[INFO] 按 Ctrl+C 停止服务")
    print("")

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n[INFO] 服务已停止")
        server.server_close()


if __name__ == "__main__":
    main()
